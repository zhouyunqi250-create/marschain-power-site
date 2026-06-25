#!/usr/bin/env python3
"""Build a best-effort MarsChain power ranking from public explorer APIs.

The explorer exposes per-address power but not a public leaderboard endpoint.
This script samples recent blocks and transactions, extracts candidate
addresses, queries each address's power, and ranks addresses by discovered
power. The result is a "discovered addresses" leaderboard, not a guaranteed
full-network ranking.
"""

from __future__ import annotations

import argparse
import concurrent.futures
import csv
import json
import math
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any


BASE_URL = "https://explorer.marschain.net/api"
DEFAULT_RPC_URL = "https://rpcs.marschain.net"
POWER_CONTRACT_ADDRESS = "0x0000000000000000000000000000000000001001"
GET_DAILY_TOTAL_POWER_HISTORY_SELECTOR = "0x1c1b5cdf"
TOKENS_BURNED_EVENT_TOPIC = "0xccbea4088a3b7ae9ca2d15fab9a9742a4075b4d7247768a1eecea917565aba00"
DEFAULT_CACHE_TTL_SECONDS = 3 * 60 * 60
DEFAULT_ADDRESS_POOL_FILE = "output/marschain_address_pool.json"
BEIJING_OFFSET_SECONDS = 8 * 60 * 60
STATISTICS_DAY_START_HOUR = 8
STATISTICS_DAY_START_LABEL = f"{STATISTICS_DAY_START_HOUR:02d}:00"
MARS_TOTAL_SUPPLY_CAP_TOKENS = 200_000_000_000
MARS_INITIAL_CYCLE_OUTPUT_TOKENS = 100_000_000_000
MARS_HALVING_PERIOD_DAYS = 448
MARS_MINER_SHARE = 0.75
MARS_NODE_SHARE = 0.25
DEFAULT_HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Accept": "application/json, text/plain, */*",
    "Referer": "https://explorer.marschain.net/",
}
RPC_HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Accept": "application/json",
    "Content-Type": "application/json",
}


@dataclass
class RankedAddress:
    address: str
    power: int
    total_burned_amount: int
    burned_amount: int
    tx_seen: int
    miner_seen: int
    log_seen: int
    upline_seen: int
    source_score: int
    upline1: str | None
    upline2: str | None
    nodes_count: int | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "address": self.address,
            "power": str(self.power),
            "power_display": format_units(self.power),
            "total_burned_amount": str(self.total_burned_amount),
            "total_burned_amount_display": format_token(self.total_burned_amount),
            "burned_amount": str(self.burned_amount),
            "burned_amount_display": format_token(self.burned_amount),
            "tx_seen": self.tx_seen,
            "miner_seen": self.miner_seen,
            "log_seen": self.log_seen,
            "upline_seen": self.upline_seen,
            "source_score": self.source_score,
            "upline1": self.upline1,
            "upline2": self.upline2,
            "nodes_count": self.nodes_count,
        }


def request_json(path: str, params: dict[str, Any] | None = None, retries: int = 6) -> Any:
    if path.startswith("http://") or path.startswith("https://"):
        url = path
    else:
        url = BASE_URL + path
    if params:
        url = f"{url}?{urllib.parse.urlencode(params)}"

    last_error: Exception | None = None
    for attempt in range(1, retries + 1):
        req = urllib.request.Request(url, headers=DEFAULT_HEADERS)
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                return json.loads(resp.read().decode("utf-8"))
        except urllib.error.HTTPError as exc:
            body = exc.read().decode("utf-8", "replace")
            last_error = RuntimeError(f"HTTP {exc.code} for {url}: {body[:200]}")
            if exc.code in {429, 500, 502, 503, 504} and attempt < retries:
                time.sleep(min(10, 0.8 * attempt * attempt))
                continue
            raise last_error
        except Exception as exc:  # pragma: no cover - network variability
            last_error = exc
            if attempt < retries:
                time.sleep(min(10, 0.8 * attempt * attempt))
                continue
            raise
    if last_error:
        raise last_error
    raise RuntimeError(f"Request failed for {url}")


def rpc_json(rpc_url: str, payload: dict[str, Any] | list[dict[str, Any]], retries: int = 5) -> Any:
    data = json.dumps(payload).encode("utf-8")
    last_error: Exception | None = None
    for attempt in range(1, retries + 1):
        req = urllib.request.Request(rpc_url, data=data, headers=RPC_HEADERS, method="POST")
        try:
            with urllib.request.urlopen(req, timeout=45) as resp:
                parsed = json.loads(resp.read().decode("utf-8"))
            if isinstance(parsed, dict) and parsed.get("error"):
                raise RuntimeError(f"RPC error for {parsed.get('id')}: {parsed['error']}")
            return parsed
        except Exception as exc:  # pragma: no cover - public RPC variability
            last_error = exc
            if attempt < retries:
                time.sleep(min(10, 0.8 * attempt * attempt))
                continue
            raise
    if last_error:
        raise last_error
    raise RuntimeError(f"RPC request failed for {rpc_url}")


def rpc_call(rpc_url: str, method: str, params: list[Any] | None = None) -> Any:
    payload = {"jsonrpc": "2.0", "id": 1, "method": method, "params": params or []}
    parsed = rpc_json(rpc_url, payload)
    if not isinstance(parsed, dict):
        raise RuntimeError(f"Unexpected RPC response for {method}: {type(parsed).__name__}")
    if parsed.get("error"):
        raise RuntimeError(f"RPC error for {method}: {parsed['error']}")
    return parsed.get("result")


def eth_call(rpc_url: str, to_address: str, data: str) -> str:
    result = rpc_call(
        rpc_url,
        "eth_call",
        [{"to": to_address, "data": data}, "latest"],
    )
    if not isinstance(result, str) or not result.startswith("0x"):
        raise RuntimeError(f"Unexpected eth_call response: {type(result).__name__}")
    return result


def get_block_timestamp(rpc_url: str, block_number: int) -> int | None:
    block = rpc_call(rpc_url, "eth_getBlockByNumber", [hex(block_number), False])
    if not isinstance(block, dict):
        return None
    return hex_to_int(block.get("timestamp"))


def find_first_block_at_or_after_timestamp(
    rpc_url: str,
    low_block: int,
    high_block: int,
    timestamp: int,
) -> int | None:
    left = max(0, low_block)
    right = max(left, high_block)
    found: int | None = None
    while left <= right:
        mid = (left + right) // 2
        block_ts = get_block_timestamp(rpc_url, mid)
        if block_ts is None:
            return None
        if block_ts >= timestamp:
            found = mid
            right = mid - 1
        else:
            left = mid + 1
    return found


def decode_uint256_words(data: str) -> list[int]:
    raw = data[2:] if data.startswith("0x") else data
    if len(raw) % 64 != 0:
        raise ValueError("ABI data length is not a multiple of 32 bytes")
    return [int(raw[idx : idx + 64], 16) for idx in range(0, len(raw), 64)]


def decode_uint256_array_pair(data: str) -> tuple[list[int], list[int]]:
    words = decode_uint256_words(data)
    if len(words) < 4:
        raise ValueError("ABI data is too short for two dynamic arrays")

    def read_array(offset_bytes: int) -> list[int]:
        start = offset_bytes // 32
        if start >= len(words):
            raise ValueError("ABI array offset is out of range")
        length = words[start]
        end = start + 1 + length
        if end > len(words):
            raise ValueError("ABI array length is out of range")
        return words[start + 1 : end]

    return read_array(words[0]), read_array(words[1])


def fetch_daily_total_power_history(rpc_url: str) -> tuple[list[int], list[int]]:
    data = eth_call(rpc_url, POWER_CONTRACT_ADDRESS, GET_DAILY_TOTAL_POWER_HISTORY_SELECTOR)
    days, powers = decode_uint256_array_pair(data)
    if len(days) != len(powers):
        raise ValueError("Daily total power history arrays have different lengths")
    return days, powers


def format_beijing_datetime(timestamp: int) -> str:
    return time.strftime("%Y-%m-%d %H:%M:%S", time.gmtime(timestamp + BEIJING_OFFSET_SECONDS))


def format_beijing_date(timestamp: int) -> str:
    return time.strftime("%Y-%m-%d", time.gmtime(timestamp + BEIJING_OFFSET_SECONDS))


def build_statistics_window_meta(reference_timestamp: int) -> dict[str, Any]:
    # Use the latest completed Beijing statistics window. Beijing 08:00 is UTC 00:00,
    # which matches the project day boundary.
    day_start_seconds = STATISTICS_DAY_START_HOUR * 60 * 60
    end_timestamp = (
        ((reference_timestamp + BEIJING_OFFSET_SECONDS - day_start_seconds) // 86_400) * 86_400
        - BEIJING_OFFSET_SECONDS
        + day_start_seconds
    )
    start_timestamp = max(0, end_timestamp - 86_400)
    start_local = format_beijing_datetime(start_timestamp)
    end_local = format_beijing_datetime(end_timestamp)
    return {
        "statistics_timezone": "Asia/Shanghai",
        "statistics_timezone_label": "北京时间",
        "statistics_day_start_hour": STATISTICS_DAY_START_HOUR,
        "statistics_window_start_timestamp": start_timestamp,
        "statistics_window_end_timestamp": end_timestamp,
        "statistics_window_start_local": start_local,
        "statistics_window_end_local": end_local,
        "statistics_window_label": f"{start_local} 至 {end_local}",
        "statistics_day_label": format_beijing_date(start_timestamp),
    }


def format_chinese_amount(value: float, decimals: int = 3) -> str:
    if value >= 1e12:
        return f"{value / 1e12:.{decimals}f}万亿"
    if value >= 1e8:
        return f"{value / 1e8:.{decimals}f}亿"
    if value >= 1e4:
        return f"{value / 1e4:.{decimals}f}万"
    if float(value).is_integer():
        return str(int(value))
    return f"{value:.{decimals}f}"


def format_yi_tokens(value: float, decimals: int = 3) -> str:
    return f"{value / 1e8:.{decimals}f}亿"


def build_mars_emission_meta(
    rpc_url: str,
    reference_timestamp: int,
    network_total_power: int,
) -> dict[str, Any]:
    genesis_timestamp: int | None = None
    emission_error: str | None = None
    try:
        genesis_timestamp = get_block_timestamp(rpc_url, 0)
    except Exception as exc:  # pragma: no cover - public RPC variability
        emission_error = str(exc)

    elapsed_days = 0
    if genesis_timestamp is not None and reference_timestamp >= genesis_timestamp:
        elapsed_days = max(0, (reference_timestamp - genesis_timestamp) // 86_400)
    current_cycle = int(elapsed_days // MARS_HALVING_PERIOD_DAYS) + 1
    cycle_output_tokens = MARS_INITIAL_CYCLE_OUTPUT_TOKENS * math.pow(0.5, current_cycle - 1)
    daily_total_tokens = cycle_output_tokens / MARS_HALVING_PERIOD_DAYS
    daily_miner_tokens = daily_total_tokens * MARS_MINER_SHARE
    daily_node_tokens = daily_total_tokens * MARS_NODE_SHARE
    power_required = network_total_power / daily_miner_tokens if daily_miner_tokens and network_total_power else None

    meta = {
        "emission_basis": "MarsChain official model: total supply 2000亿, halving every 448 days, miner 75%, node 25%",
        "emission_reference_timestamp": reference_timestamp,
        "emission_genesis_timestamp": genesis_timestamp,
        "emission_current_cycle": current_cycle,
        "emission_cycle_elapsed_days": int(elapsed_days % MARS_HALVING_PERIOD_DAYS),
        "emission_total_supply_cap_tokens": MARS_TOTAL_SUPPLY_CAP_TOKENS,
        "emission_total_supply_cap_display": "2000亿",
        "emission_initial_cycle_output_tokens": MARS_INITIAL_CYCLE_OUTPUT_TOKENS,
        "emission_halving_period_days": MARS_HALVING_PERIOD_DAYS,
        "emission_miner_share": MARS_MINER_SHARE,
        "emission_node_share": MARS_NODE_SHARE,
        "emission_cycle_output_tokens": cycle_output_tokens,
        "emission_daily_total_tokens": daily_total_tokens,
        "emission_daily_total_display": f"{format_yi_tokens(daily_total_tokens)}/日",
        "emission_daily_miner_tokens": daily_miner_tokens,
        "emission_daily_miner_display": f"{format_yi_tokens(daily_miner_tokens)}/日",
        "emission_daily_node_tokens": daily_node_tokens,
        "emission_daily_node_display": f"{format_yi_tokens(daily_node_tokens)}/日",
        "power_required_per_mars_daily": power_required,
        "power_required_per_mars_daily_display": format_chinese_amount(power_required) if power_required is not None else None,
    }
    if emission_error:
        meta["emission_error"] = emission_error
    return meta


def normalize_address(value: str | None) -> str | None:
    if not value or not isinstance(value, str):
        return None
    value = value.strip()
    if not value.startswith("0x") or len(value) != 42:
        return None
    return value.lower()


def is_probably_user_address(value: str | None) -> bool:
    address = normalize_address(value)
    if not address:
        return False
    lower = address.lower()
    if lower == "0x0000000000000000000000000000000000000000":
        return False
    # Low numeric values can appear in event topics as uint256 token IDs.
    # Real externally-owned accounts in this range are vanishingly unlikely.
    if lower.startswith("0x000000000000000000000000"):
        return False
    return True


def topic_to_probable_address(topic: str | None) -> str | None:
    if not isinstance(topic, str) or not topic.startswith("0x") or len(topic) != 66:
        return None
    # Solidity encodes indexed address topics as 12 zero bytes + 20-byte address.
    if topic[2:26] != "0" * 24:
        return None
    return normalize_address("0x" + topic[-40:])


def hex_to_int(value: str | int | None) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, str):
        try:
            return int(value, 16) if value.startswith("0x") else int(value)
        except ValueError:
            return None
    return None


def decode_log_uint256_words(data: str | None) -> list[int]:
    if not isinstance(data, str):
        return []
    raw = data[2:] if data.startswith("0x") else data
    if not raw or len(raw) % 64 != 0:
        return []
    try:
        return [int(raw[idx : idx + 64], 16) for idx in range(0, len(raw), 64)]
    except ValueError:
        return []


def format_units(raw: int) -> str:
    return format_chinese_amount(raw)


def format_token(raw: int) -> str:
    return f"{raw / 10**18:.6f}"


def format_token_chinese(raw: int | None) -> str | None:
    if raw is None:
        return None
    return format_chinese_amount(raw / 10**18)


def format_price(value: object) -> str | None:
    if value is None or value == "":
        return None
    try:
        number = float(str(value))
    except (TypeError, ValueError):
        return str(value)
    if number < 1:
        return f"{number:.6f}"
    return f"{number:.3f}"


def format_percent(value: float) -> str:
    return f"{value * 100:.2f}%"


def load_cache(path: Path | None) -> dict[str, dict[str, Any]]:
    if not path or not path.exists():
        return {}
    try:
        data = json.loads(path.read_text())
    except Exception:
        return {}
    if not isinstance(data, dict):
        return {}
    normalized: dict[str, dict[str, Any]] = {}
    for key, value in data.items():
        address = normalize_address(str(key))
        if address and isinstance(value, dict):
            normalized[address] = value
    return normalized


def save_cache(path: Path | None, cache: dict[str, dict[str, Any]]) -> None:
    if not path:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(cache, ensure_ascii=False, indent=2) + "\n")


def load_address_pool(path: Path | None) -> tuple[dict[str, dict[str, Any]], dict[str, Any]]:
    if not path or not path.exists():
        return {}, {}
    try:
        data = json.loads(path.read_text())
    except Exception:
        return {}, {}
    if isinstance(data, dict) and isinstance(data.get("addresses"), dict):
        raw_addresses = data.get("addresses") or {}
        raw_meta = data.get("meta") if isinstance(data.get("meta"), dict) else {}
    elif isinstance(data, dict):
        raw_addresses = data
        raw_meta = {}
    else:
        return {}, {}

    normalized: dict[str, dict[str, Any]] = {}
    for key, value in raw_addresses.items():
        address = normalize_address(str(key))
        if address and isinstance(value, dict):
            normalized[address] = dict(value)
    return normalized, dict(raw_meta)


def save_address_pool(path: Path | None, pool: dict[str, dict[str, Any]], meta: dict[str, Any] | None = None) -> None:
    if not path:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(
            {
                "meta": meta or {},
                "addresses": pool,
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n"
    )


def ensure_address_pool_entry(
    pool: dict[str, dict[str, Any]],
    address: str,
    observed_at: int,
    source: str,
) -> dict[str, Any]:
    entry = pool.get(address)
    if entry is None:
        entry = {
            "first_seen_at": observed_at,
            "first_seen_source": source,
            "last_seen_at": observed_at,
            "last_seen_source": source,
            "tx_seen_total": 0,
            "miner_seen_total": 0,
            "log_seen_total": 0,
            "upline_seen_total": 0,
            "source_score_total": 0,
            "discovery_rounds": 0,
            "last_network_refresh_at": None,
            "last_network_refresh_source": None,
            "network_refresh_count": 0,
            "first_seen_block": None,
            "last_seen_block": None,
        }
        pool[address] = entry
    else:
        if entry.get("first_seen_at") is None:
            entry["first_seen_at"] = observed_at
        if entry.get("first_seen_source") is None:
            entry["first_seen_source"] = source
        if entry.get("last_seen_at") is None or int(observed_at) >= int(entry.get("last_seen_at") or 0):
            entry["last_seen_at"] = observed_at
            entry["last_seen_source"] = source
    return entry


def merge_address_pool_counters(
    pool: dict[str, dict[str, Any]],
    tx_counter: Counter[str],
    miner_counter: Counter[str],
    log_counter: Counter[str],
    upline_counter: Counter[str],
    observed_at: int,
) -> int:
    new_addresses = 0
    source_specs = (
        ("tx", tx_counter, "tx_seen_total", 10),
        ("miner", miner_counter, "miner_seen_total", 20),
        ("log", log_counter, "log_seen_total", 15),
        ("upline", upline_counter, "upline_seen_total", 5),
    )
    for source, counter, total_key, weight in source_specs:
        for address, count in counter.items():
            if count <= 0:
                continue
            entry = pool.get(address)
            if entry is None:
                new_addresses += 1
            entry = ensure_address_pool_entry(pool, address, observed_at, source)
            entry[total_key] = int(entry.get(total_key, 0) or 0) + int(count)
            entry["source_score_total"] = int(entry.get("source_score_total", 0) or 0) + int(count) * weight
            entry["discovery_rounds"] = int(entry.get("discovery_rounds", 0) or 0) + 1
    return new_addresses


def add_cache_addresses_to_pool(
    pool: dict[str, dict[str, Any]],
    cache: dict[str, dict[str, Any]],
    observed_at: int,
) -> int:
    added = 0
    for address, payload in cache.items():
        if address in pool:
            continue
        entry = ensure_address_pool_entry(pool, address, observed_at, "cache")
        cached_at = payload.get("cached_at")
        if isinstance(cached_at, int | float):
            entry["last_network_refresh_at"] = int(cached_at)
            entry["last_network_refresh_source"] = "cache"
            entry["network_refresh_count"] = max(1, int(entry.get("network_refresh_count", 0) or 0))
        added += 1
    return added


def mark_network_refresh(
    pool: dict[str, dict[str, Any]],
    address: str,
    observed_at: int,
    source: str,
) -> None:
    entry = pool.get(address)
    if entry is None:
        entry = ensure_address_pool_entry(pool, address, observed_at, source)
    entry["last_network_refresh_at"] = observed_at
    entry["last_network_refresh_source"] = source
    entry["network_refresh_count"] = int(entry.get("network_refresh_count", 0) or 0) + 1


def is_cache_entry_fresh(payload: dict[str, Any], ttl_seconds: int | None) -> bool:
    if ttl_seconds is None or ttl_seconds < 0:
        return True
    if ttl_seconds == 0:
        return False
    cached_at = payload.get("cached_at")
    if not isinstance(cached_at, int | float):
        return False
    return int(time.time()) - int(cached_at) <= ttl_seconds


def collect_candidates(
    tx_pages: int,
    tx_limit: int,
    block_pages: int,
    block_limit: int,
    include_to: bool,
    workers: int,
    progress: bool,
) -> tuple[Counter[str], Counter[str]]:
    tx_counter: Counter[str] = Counter()
    miner_counter: Counter[str] = Counter()

    def fetch_tx_page(page: int) -> tuple[int, dict[str, Any]]:
        payload = request_json("/transactions", {"page": page, "limit": tx_limit})
        return page, payload

    def fetch_block_page(page: int) -> tuple[int, dict[str, Any]]:
        payload = request_json("/blocks", {"page": page, "limit": block_limit})
        return page, payload

    tx_workers = max(1, min(workers, 32))
    block_workers = max(1, min(workers, 16))

    if tx_pages:
        with concurrent.futures.ThreadPoolExecutor(max_workers=tx_workers) as pool:
            future_map = {pool.submit(fetch_tx_page, page): page for page in range(1, tx_pages + 1)}
            for idx, future in enumerate(concurrent.futures.as_completed(future_map), start=1):
                payload = future.result()[1]
                for tx in payload.get("transactions", []):
                    sender = normalize_address(tx.get("from"))
                    if is_probably_user_address(sender):
                        tx_counter[sender] += 1
                    if include_to:
                        receiver = normalize_address(tx.get("to"))
                        if is_probably_user_address(receiver):
                            tx_counter[receiver] += 1
                if progress and idx % 100 == 0:
                    print(f"[info] tx-page scan: {idx}/{tx_pages}", file=sys.stderr)

    if block_pages:
        with concurrent.futures.ThreadPoolExecutor(max_workers=block_workers) as pool:
            future_map = {pool.submit(fetch_block_page, page): page for page in range(1, block_pages + 1)}
            for idx, future in enumerate(concurrent.futures.as_completed(future_map), start=1):
                payload = future.result()[1]
                for block in payload.get("blocks", []):
                    miner = normalize_address(block.get("miner"))
                    if is_probably_user_address(miner):
                        miner_counter[miner] += 1
                if progress and idx % 100 == 0:
                    print(f"[info] block-page scan: {idx}/{block_pages}", file=sys.stderr)

    return tx_counter, miner_counter


def collect_rpc_block_candidates(
    rpc_url: str,
    rpc_blocks: int,
    rpc_start_block: int | None,
    rpc_batch_size: int,
    rpc_workers: int,
    include_to: bool,
    progress: bool,
) -> tuple[Counter[str], Counter[str], dict[str, Any]]:
    tx_counter: Counter[str] = Counter()
    miner_counter: Counter[str] = Counter()
    meta: dict[str, Any] = {
        "rpc_url": rpc_url,
        "rpc_blocks_requested": rpc_blocks,
        "rpc_blocks_scanned": 0,
        "rpc_transactions_seen": 0,
        "rpc_error_count": 0,
    }
    if not rpc_url or rpc_blocks <= 0:
        return tx_counter, miner_counter, meta

    try:
        latest_block = hex_to_int(rpc_call(rpc_url, "eth_blockNumber"))
    except Exception as exc:
        meta["rpc_error"] = str(exc)
        print(f"[warn] rpc latest block lookup failed: {exc}", file=sys.stderr)
        return tx_counter, miner_counter, meta

    if latest_block is None:
        meta["rpc_error"] = "eth_blockNumber returned an invalid value"
        return tx_counter, miner_counter, meta

    high_block = latest_block if rpc_start_block is None else min(rpc_start_block, latest_block)
    low_block = max(0, high_block - rpc_blocks + 1)
    batch_size = max(1, min(rpc_batch_size, 200))
    meta.update(
        {
            "rpc_latest_block": latest_block,
            "rpc_start_block": high_block,
            "rpc_end_block": low_block,
            "rpc_batch_size": batch_size,
            "rpc_workers": rpc_workers,
        }
    )

    batches: list[list[int]] = []
    current = high_block
    while current >= low_block:
        end = max(low_block, current - batch_size + 1)
        batches.append(list(range(current, end - 1, -1)))
        current = end - 1

    def fetch_batch(numbers: list[int]) -> tuple[Counter[str], Counter[str], int, int]:
        payload = [
            {
                "jsonrpc": "2.0",
                "id": number,
                "method": "eth_getBlockByNumber",
                "params": [hex(number), True],
            }
            for number in numbers
        ]
        parsed = rpc_json(rpc_url, payload)
        if not isinstance(parsed, list):
            raise RuntimeError(f"Unexpected RPC batch response: {type(parsed).__name__}")

        local_tx_counter: Counter[str] = Counter()
        local_miner_counter: Counter[str] = Counter()
        scanned = 0
        tx_seen = 0
        for item in parsed:
            if not isinstance(item, dict) or item.get("error"):
                continue
            block = item.get("result")
            if not isinstance(block, dict):
                continue
            scanned += 1
            miner = normalize_address(block.get("miner") or block.get("author"))
            if is_probably_user_address(miner):
                local_miner_counter[miner] += 1
            for tx in block.get("transactions", []):
                if not isinstance(tx, dict):
                    continue
                tx_seen += 1
                sender = normalize_address(tx.get("from"))
                if is_probably_user_address(sender):
                    local_tx_counter[sender] += 1
                if include_to:
                    receiver = normalize_address(tx.get("to"))
                    if is_probably_user_address(receiver):
                        local_tx_counter[receiver] += 1
        return local_tx_counter, local_miner_counter, scanned, tx_seen

    with concurrent.futures.ThreadPoolExecutor(max_workers=max(1, min(rpc_workers, 12))) as pool:
        future_map = {pool.submit(fetch_batch, batch): batch for batch in batches}
        for idx, future in enumerate(concurrent.futures.as_completed(future_map), start=1):
            try:
                batch_tx, batch_miner, scanned, tx_seen = future.result()
            except Exception as exc:
                meta["rpc_error_count"] += 1
                if meta["rpc_error_count"] <= 10:
                    print(f"[warn] rpc block batch failed: {exc}", file=sys.stderr)
                continue
            tx_counter.update(batch_tx)
            miner_counter.update(batch_miner)
            meta["rpc_blocks_scanned"] += scanned
            meta["rpc_transactions_seen"] += tx_seen
            if progress and idx % 50 == 0:
                print(
                    f"[info] rpc block scan: {idx}/{len(batches)} batches, "
                    f"{meta['rpc_blocks_scanned']}/{rpc_blocks} blocks",
                    file=sys.stderr,
                )

    return tx_counter, miner_counter, meta


def collect_statistics_window_active_addresses(
    rpc_url: str,
    latest_block: int | None,
    start_timestamp: int,
    end_timestamp: int,
    rpc_batch_size: int,
    rpc_workers: int,
    progress: bool,
) -> dict[str, Any]:
    meta: dict[str, Any] = {
        "statistics_window_active_address_basis": f"unique transaction sender/receiver wallet addresses in the latest completed Beijing {STATISTICS_DAY_START_LABEL} statistics window",
        "statistics_window_active_wallet_address_count": None,
        "statistics_window_active_blocks_scanned": 0,
        "statistics_window_active_transactions_seen": 0,
        "statistics_window_transaction_volume_wei": None,
        "statistics_window_transaction_volume_display": None,
        "statistics_window_active_error_count": 0,
    }
    if not rpc_url:
        meta["statistics_window_active_error"] = "missing rpc url"
        return meta

    try:
        if latest_block is None:
            latest_block = hex_to_int(rpc_call(rpc_url, "eth_blockNumber"))
        if latest_block is None:
            raise RuntimeError("eth_blockNumber returned an invalid value")
        start_block = find_first_block_at_or_after_timestamp(rpc_url, 0, latest_block, start_timestamp)
        end_marker_block = find_first_block_at_or_after_timestamp(rpc_url, 0, latest_block, end_timestamp)
    except Exception as exc:
        meta["statistics_window_active_error"] = str(exc)
        print(f"[warn] statistics window active address lookup failed: {exc}", file=sys.stderr)
        return meta

    if start_block is None:
        meta["statistics_window_active_error"] = "window start block not found"
        return meta
    if end_marker_block is None:
        end_marker_block = latest_block + 1

    low_block = start_block
    high_block = min(latest_block, end_marker_block - 1)
    meta.update(
        {
            "statistics_window_active_start_block": low_block,
            "statistics_window_active_end_block": high_block,
            "statistics_window_active_end_marker_block": end_marker_block,
        }
    )
    if high_block < low_block:
        meta["statistics_window_active_wallet_address_count"] = 0
        return meta

    batch_size = max(1, min(rpc_batch_size, 200))
    batches: list[list[int]] = []
    current = high_block
    while current >= low_block:
        end = max(low_block, current - batch_size + 1)
        batches.append(list(range(current, end - 1, -1)))
        current = end - 1

    def fetch_batch(numbers: list[int]) -> tuple[set[str], int, int, int]:
        payload = [
            {
                "jsonrpc": "2.0",
                "id": number,
                "method": "eth_getBlockByNumber",
                "params": [hex(number), True],
            }
            for number in numbers
        ]
        parsed = rpc_json(rpc_url, payload)
        if not isinstance(parsed, list):
            raise RuntimeError(f"Unexpected RPC batch response: {type(parsed).__name__}")

        local_addresses: set[str] = set()
        scanned = 0
        tx_seen = 0
        tx_volume_wei = 0
        for item in parsed:
            if not isinstance(item, dict) or item.get("error"):
                continue
            block = item.get("result")
            if not isinstance(block, dict):
                continue
            scanned += 1
            for tx in block.get("transactions", []):
                if not isinstance(tx, dict):
                    continue
                tx_seen += 1
                value_wei = hex_to_int(tx.get("value"))
                if value_wei and value_wei > 0:
                    tx_volume_wei += value_wei
                sender = normalize_address(tx.get("from"))
                if is_probably_user_address(sender):
                    local_addresses.add(sender)
                receiver = normalize_address(tx.get("to"))
                if is_probably_user_address(receiver):
                    local_addresses.add(receiver)
        return local_addresses, scanned, tx_seen, tx_volume_wei

    active_addresses: set[str] = set()
    total_transaction_volume_wei = 0
    with concurrent.futures.ThreadPoolExecutor(max_workers=max(1, min(rpc_workers, 12))) as pool:
        future_map = {pool.submit(fetch_batch, batch): batch for batch in batches}
        for idx, future in enumerate(concurrent.futures.as_completed(future_map), start=1):
            block_range = future_map[future]
            try:
                batch_addresses, scanned, tx_seen, tx_volume_wei = future.result()
            except Exception as exc:
                meta["statistics_window_active_error_count"] += 1
                if meta["statistics_window_active_error_count"] <= 10:
                    print(
                        f"[warn] statistics window active block batch {block_range[-1]}-{block_range[0]} failed: {exc}",
                        file=sys.stderr,
                    )
                continue
            active_addresses.update(batch_addresses)
            meta["statistics_window_active_blocks_scanned"] += scanned
            meta["statistics_window_active_transactions_seen"] += tx_seen
            total_transaction_volume_wei += tx_volume_wei
            if progress and idx % 50 == 0:
                print(
                    f"[info] statistics window active scan: {idx}/{len(batches)} batches, "
                    f"{meta['statistics_window_active_blocks_scanned']} blocks",
                    file=sys.stderr,
                )

    meta["statistics_window_active_wallet_address_count"] = len(active_addresses)
    meta["statistics_window_transaction_volume_wei"] = total_transaction_volume_wei
    meta["statistics_window_transaction_volume_display"] = format_token_chinese(total_transaction_volume_wei)
    meta["statistics_window_transaction_volume_basis"] = f"sum native MARS transaction value in the latest completed Beijing {STATISTICS_DAY_START_LABEL} statistics window"
    return meta


def collect_rpc_log_candidates(
    rpc_url: str,
    rpc_log_blocks: int,
    rpc_log_start_block: int | None,
    rpc_log_chunk_size: int,
    rpc_log_workers: int,
    progress: bool,
) -> tuple[Counter[str], dict[str, Any]]:
    log_counter: Counter[str] = Counter()
    meta: dict[str, Any] = {
        "rpc_log_url": rpc_url,
        "rpc_log_contract": POWER_CONTRACT_ADDRESS,
        "rpc_log_blocks_requested": rpc_log_blocks,
        "rpc_log_blocks_scanned": 0,
        "rpc_logs_seen": 0,
        "rpc_log_error_count": 0,
    }
    if not rpc_url or rpc_log_blocks <= 0:
        return log_counter, meta

    try:
        latest_block = hex_to_int(rpc_call(rpc_url, "eth_blockNumber"))
    except Exception as exc:
        meta["rpc_log_error"] = str(exc)
        print(f"[warn] rpc log latest block lookup failed: {exc}", file=sys.stderr)
        return log_counter, meta

    if latest_block is None:
        meta["rpc_log_error"] = "eth_blockNumber returned an invalid value"
        return log_counter, meta

    latest_timestamp = get_block_timestamp(rpc_url, latest_block)
    statistics_window_start_timestamp: int | None = None
    statistics_window_end_timestamp: int | None = None
    statistics_window_start_block: int | None = None
    statistics_window_end_block: int | None = None
    period_windows: dict[int, dict[str, int | str | bool | None]] = {}
    if latest_timestamp is not None:
        statistics_window_meta = build_statistics_window_meta(latest_timestamp)
        statistics_window_start_timestamp = statistics_window_meta["statistics_window_start_timestamp"]
        statistics_window_end_timestamp = statistics_window_meta["statistics_window_end_timestamp"]
        try:
            statistics_window_start_block = find_first_block_at_or_after_timestamp(
                rpc_url,
                0,
                latest_block,
                statistics_window_start_timestamp,
            )
            statistics_window_end_block = find_first_block_at_or_after_timestamp(
                rpc_url,
                0,
                latest_block,
                statistics_window_end_timestamp,
            )
            for days in (7, 30):
                period_start_timestamp = max(0, statistics_window_end_timestamp - days * 86_400)
                period_start_block = find_first_block_at_or_after_timestamp(
                    rpc_url,
                    0,
                    latest_block,
                    period_start_timestamp,
                )
                period_windows[days] = {
                    "start_timestamp": period_start_timestamp,
                    "end_timestamp": statistics_window_end_timestamp,
                    "start_local": format_beijing_datetime(period_start_timestamp),
                    "end_local": format_beijing_datetime(statistics_window_end_timestamp),
                    "label": f"{format_beijing_datetime(period_start_timestamp)} 至 {format_beijing_datetime(statistics_window_end_timestamp)}",
                    "start_block": period_start_block,
                    "end_block": statistics_window_end_block,
                }
        except Exception as exc:
            meta["rpc_log_statistics_window_error"] = str(exc)
            statistics_window_start_block = None
            statistics_window_end_block = None

    high_block = latest_block if rpc_log_start_block is None else min(rpc_log_start_block, latest_block)
    low_block = max(0, high_block - rpc_log_blocks + 1)
    effective_blocks = high_block - low_block + 1
    chunk_size = max(1, min(rpc_log_chunk_size, 100_000))
    meta.update(
        {
            "rpc_log_latest_block": latest_block,
            "rpc_log_start_block": high_block,
            "rpc_log_end_block": low_block,
            "rpc_log_blocks_effective": effective_blocks,
            "rpc_log_latest_timestamp": latest_timestamp,
            "rpc_log_statistics_window_start_timestamp": statistics_window_start_timestamp,
            "rpc_log_statistics_window_end_timestamp": statistics_window_end_timestamp,
            "rpc_log_statistics_window_start_block": statistics_window_start_block,
            "rpc_log_statistics_window_end_block": statistics_window_end_block,
            "rpc_log_today_start_timestamp": statistics_window_start_timestamp,
            "rpc_log_today_start_block": statistics_window_start_block,
            "rpc_log_chunk_size": chunk_size,
            "rpc_log_workers": rpc_log_workers,
        }
    )

    ranges: list[tuple[int, int]] = []
    current = high_block
    while current >= low_block:
        start = max(low_block, current - chunk_size + 1)
        ranges.append((start, current))
        current = start - 1

    first_seen_block: dict[str, int] = {}
    burned_by_statistics_window = 0
    burned_by_period: dict[int, int] = {7: 0, 30: 0}

    def fetch_range(block_range: tuple[int, int]) -> tuple[Counter[str], dict[str, int], int, dict[int, int], int, int]:
        start, end = block_range
        parsed = rpc_call(
            rpc_url,
            "eth_getLogs",
            [
                {
                    "fromBlock": hex(start),
                    "toBlock": hex(end),
                    "address": POWER_CONTRACT_ADDRESS,
                }
            ],
        )
        if not isinstance(parsed, list):
            raise RuntimeError(f"Unexpected eth_getLogs response: {type(parsed).__name__}")

        local_counter: Counter[str] = Counter()
        local_first_seen: dict[str, int] = {}
        local_burned_by_statistics_window = 0
        local_burned_by_period: dict[int, int] = {7: 0, 30: 0}
        for log in parsed:
            if not isinstance(log, dict):
                continue
            block_number = hex_to_int(log.get("blockNumber"))
            topics = log.get("topics")
            if not isinstance(topics, list):
                continue
            if topics and str(topics[0]).lower() == TOKENS_BURNED_EVENT_TOPIC:
                words = decode_log_uint256_words(log.get("data"))
                burned_amount = words[0] if words else 0
                if block_number is not None and burned_amount > 0:
                    if (
                        isinstance(statistics_window_start_block, int)
                        and isinstance(statistics_window_end_block, int)
                        and statistics_window_start_block <= block_number < statistics_window_end_block
                    ):
                        local_burned_by_statistics_window += burned_amount
                    for days, window in period_windows.items():
                        start_block = window.get("start_block")
                        end_block = window.get("end_block")
                        if (
                            isinstance(start_block, int)
                            and isinstance(end_block, int)
                            and start_block <= block_number < end_block
                        ):
                            local_burned_by_period[days] = local_burned_by_period.get(days, 0) + burned_amount
            for topic in topics[1:]:
                address = topic_to_probable_address(topic)
                if is_probably_user_address(address):
                    local_counter[address] += 1
                    if block_number is not None:
                        previous = local_first_seen.get(address)
                        if previous is None or block_number < previous:
                            local_first_seen[address] = block_number
        return local_counter, local_first_seen, local_burned_by_statistics_window, local_burned_by_period, end - start + 1, len(parsed)

    with concurrent.futures.ThreadPoolExecutor(max_workers=max(1, min(rpc_log_workers, 6))) as pool:
        future_map = {pool.submit(fetch_range, block_range): block_range for block_range in ranges}
        for idx, future in enumerate(concurrent.futures.as_completed(future_map), start=1):
            block_range = future_map[future]
            try:
                (
                    batch_counter,
                    batch_first_seen,
                    batch_burned_by_statistics_window,
                    batch_burned_by_period,
                    blocks_scanned,
                    logs_seen,
                ) = future.result()
            except Exception as exc:
                meta["rpc_log_error_count"] += 1
                if meta["rpc_log_error_count"] <= 10:
                    print(f"[warn] rpc log range {block_range[0]}-{block_range[1]} failed: {exc}", file=sys.stderr)
                continue
            log_counter.update(batch_counter)
            for address, block_number in batch_first_seen.items():
                previous = first_seen_block.get(address)
                if previous is None or block_number < previous:
                    first_seen_block[address] = block_number
            burned_by_statistics_window += batch_burned_by_statistics_window
            for days, burned_amount in batch_burned_by_period.items():
                burned_by_period[days] = burned_by_period.get(days, 0) + burned_amount
            meta["rpc_log_blocks_scanned"] += blocks_scanned
            meta["rpc_logs_seen"] += logs_seen
            if progress and idx % 5 == 0:
                print(
                    f"[info] rpc log scan: {idx}/{len(ranges)} ranges, "
                    f"{meta['rpc_log_blocks_scanned']}/{effective_blocks} blocks, "
                    f"{meta['rpc_logs_seen']} logs",
                    file=sys.stderr,
                )

    meta["rpc_log_addresses_seen"] = len(log_counter)
    meta["rpc_log_first_seen_tracked"] = len(first_seen_block)
    if statistics_window_start_block is not None and statistics_window_end_block is not None:
        meta["statistics_window_new_candidate_address_count"] = sum(
            1
            for block_number in first_seen_block.values()
            if statistics_window_start_block <= block_number < statistics_window_end_block
        )
    else:
        meta["statistics_window_new_candidate_address_count"] = None
    meta["statistics_window_new_candidate_address_basis"] = f"first POWER-contract log in the latest completed Beijing {STATISTICS_DAY_START_LABEL} statistics window"
    meta["today_new_wallet_count"] = meta["statistics_window_new_candidate_address_count"]
    meta["today_new_wallet_basis"] = meta["statistics_window_new_candidate_address_basis"]
    if statistics_window_start_block is not None and statistics_window_end_block is not None:
        meta["statistics_window_burned_tokens"] = burned_by_statistics_window
        meta["statistics_window_burned_display"] = format_token_chinese(burned_by_statistics_window)
    else:
        meta["statistics_window_burned_tokens"] = None
        meta["statistics_window_burned_display"] = None
    meta["statistics_window_burned_basis"] = f"sum TokensBurned event first uint256 amount in the latest completed Beijing {STATISTICS_DAY_START_LABEL} statistics window"
    meta["today_burned_tokens"] = meta["statistics_window_burned_tokens"]
    meta["today_burned_display"] = meta["statistics_window_burned_display"]
    meta["today_burned_basis"] = meta["statistics_window_burned_basis"]
    for days in (7, 30):
        prefix = f"period_{days}d"
        window = period_windows.get(days, {})
        start_block = window.get("start_block")
        end_block = window.get("end_block")
        meta[f"{prefix}_label"] = window.get("label")
        meta[f"{prefix}_start_timestamp"] = window.get("start_timestamp")
        meta[f"{prefix}_end_timestamp"] = window.get("end_timestamp")
        meta[f"{prefix}_start_local"] = window.get("start_local")
        meta[f"{prefix}_end_local"] = window.get("end_local")
        meta[f"{prefix}_start_block"] = start_block
        meta[f"{prefix}_end_block"] = end_block
        meta[f"{prefix}_complete"] = isinstance(start_block, int) and low_block <= start_block
        if isinstance(start_block, int) and isinstance(end_block, int):
            meta[f"{prefix}_new_candidate_address_count"] = sum(
                1
                for block_number in first_seen_block.values()
                if start_block <= block_number < end_block
            )
            burned_amount = burned_by_period.get(days, 0)
            meta[f"{prefix}_burned_tokens"] = burned_amount
            meta[f"{prefix}_burned_display"] = format_token_chinese(burned_amount)
        else:
            meta[f"{prefix}_new_candidate_address_count"] = None
            meta[f"{prefix}_burned_tokens"] = None
            meta[f"{prefix}_burned_display"] = None
        meta[f"{prefix}_new_candidate_address_basis"] = f"first POWER-contract log in the latest completed {days} Beijing statistics days"
        meta[f"{prefix}_burned_basis"] = f"sum TokensBurned event first uint256 amount in the latest completed {days} Beijing statistics days"
    return log_counter, meta


def build_row_from_payload(
    address: str,
    payload: dict[str, Any],
    tx_seen: int = 0,
    miner_seen: int = 0,
    log_seen: int = 0,
    upline_seen: int = 0,
) -> RankedAddress | None:
    power = int(payload.get("power", "0") or 0)
    total_burned_amount = int(payload.get("totalBurnedAmount", "0") or 0)
    burned_amount = int(payload.get("burnedAmount", "0") or 0)
    if power <= 0:
        return None
    return RankedAddress(
        address=address,
        power=power,
        total_burned_amount=total_burned_amount,
        burned_amount=burned_amount,
        tx_seen=tx_seen,
        miner_seen=miner_seen,
        log_seen=log_seen,
        upline_seen=upline_seen,
        source_score=tx_seen * 10 + miner_seen * 20 + log_seen * 15 + upline_seen * 5,
        upline1=normalize_address(payload.get("upline1")),
        upline2=normalize_address(payload.get("upline2")),
    )


def fetch_power_payload(
    address: str,
    cache: dict[str, dict[str, Any]],
    cache_ttl_seconds: int | None,
) -> tuple[str, dict[str, Any], str]:
    cached = cache.get(address)
    if cached is not None and is_cache_entry_fresh(cached, cache_ttl_seconds):
        return address, cached, "hit"
    try:
        payload = request_json(f"/power/{address}")
    except Exception:
        if cached is not None:
            # Keep publishing from the last known value if the public explorer API
            # has a transient failure, but do not mark the stale value as fresh.
            return address, dict(cached), "stale_fallback"
        raise
    if isinstance(payload, dict):
        payload = dict(payload)
        payload["cached_at"] = int(time.time())
    return address, payload, "refreshed"


def enrich_nodes(address: str) -> int | None:
    payload = request_json(f"/nodes/{address}")
    nodes = payload.get("nodes")
    if isinstance(nodes, list):
        return len(nodes)
    return 0 if nodes is None else None


def fetch_address_transactions_page(address: str, page: int, limit: int) -> tuple[str, int, dict[str, Any]]:
    payload = request_json(f"/addresses/{address}/transactions", {"page": page, "limit": limit})
    return address, page, payload


def collect_history_counter(
    seed_addresses: list[str],
    pages: int,
    limit: int,
    workers: int,
    progress: bool,
) -> Counter[str]:
    counter: Counter[str] = Counter()
    if not seed_addresses or pages <= 0 or limit <= 0:
        return counter

    tasks: list[tuple[str, int]] = [(address, page) for address in seed_addresses for page in range(1, pages + 1)]
    with concurrent.futures.ThreadPoolExecutor(max_workers=max(1, min(workers, 32))) as pool:
        future_map = {
            pool.submit(fetch_address_transactions_page, address, page, limit): (address, page)
            for address, page in tasks
        }
        for idx, future in enumerate(concurrent.futures.as_completed(future_map), start=1):
            address, page = future_map[future]
            try:
                _, _, payload = future.result()
            except Exception as exc:
                print(f"[warn] address tx lookup failed for {address} page {page}: {exc}", file=sys.stderr)
                continue
            for tx in payload.get("transactions", []):
                sender = normalize_address(tx.get("from"))
                receiver = normalize_address(tx.get("to"))
                if sender == address and is_probably_user_address(receiver):
                    counter[receiver] += 1
                elif receiver == address and is_probably_user_address(sender):
                    counter[sender] += 1
            if progress and idx % 100 == 0:
                print(
                    f"[info] history tx scan: {idx}/{len(tasks)} seed-pages checked",
                    file=sys.stderr,
                )
    return counter


def write_html(path: Path, rows: list[RankedAddress], meta: dict[str, Any]) -> None:
    summary_items = [
        ("生成时间", time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(meta["generated_at"]))),
        ("候选地址", f"{meta['candidate_count']:,}"),
        ("发现正算力地址", f"{meta['positive_power_count']:,}"),
        ("榜单行数", f"{meta['ranked_count']:,}"),
        ("全网总算力", format_units(meta["network_total_power"])),
        ("已发现覆盖", format_percent(meta["discovered_power_coverage"])),
        ("合约日志", f"{meta.get('rpc_logs_seen', 0):,}"),
    ]
    summary_html = "\n".join(
        f'<div class="card"><div class="label">{label}</div><div class="value">{value}</div></div>'
        for label, value in summary_items
    )
    row_html = "\n".join(
        (
            "<tr>"
            f"<td>{idx}</td>"
            f"<td class='mono'>{row.address}</td>"
            f"<td data-sort='{row.power}'>{format_units(row.power)}</td>"
            f"<td data-sort='{row.total_burned_amount}'>{format_token(row.total_burned_amount)}</td>"
            f"<td>{row.tx_seen}</td>"
            f"<td>{row.miner_seen}</td>"
            f"<td>{row.log_seen}</td>"
            f"<td>{row.upline_seen}</td>"
            f"<td>{row.nodes_count if row.nodes_count is not None else ''}</td>"
            f"<td class='mono'>{row.upline1 or ''}</td>"
            f"<td class='mono'>{row.upline2 or ''}</td>"
            "</tr>"
        )
        for idx, row in enumerate(rows, start=1)
    )
    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>MarsChain 算力排行榜</title>
  <style>
    :root {{
      --bg: #0b1020;
      --panel: #121a31;
      --panel-2: #182342;
      --line: #2b3b6b;
      --text: #e9eefc;
      --muted: #9fb0d9;
      --accent: #67d4ff;
      --accent-2: #8b5cf6;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      padding: 32px;
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      background:
        radial-gradient(circle at top right, rgba(103, 212, 255, 0.18), transparent 22%),
        linear-gradient(180deg, #0b1020 0%, #0f1730 100%);
      color: var(--text);
    }}
    .wrap {{ max-width: 1440px; margin: 0 auto; }}
    .hero {{
      background: linear-gradient(135deg, rgba(24, 35, 66, 0.98), rgba(16, 24, 48, 0.92));
      border: 1px solid var(--line);
      border-radius: 24px;
      padding: 28px 30px;
      margin-bottom: 22px;
      box-shadow: 0 24px 80px rgba(0, 0, 0, 0.35);
    }}
    .eyebrow {{
      color: var(--accent);
      font-size: 12px;
      letter-spacing: 0.16em;
      text-transform: uppercase;
      margin-bottom: 10px;
    }}
    h1 {{ margin: 0; font-size: 36px; line-height: 1.15; }}
    .sub {{
      margin-top: 10px;
      color: var(--muted);
      max-width: 920px;
      line-height: 1.6;
    }}
    .grid {{
      display: grid;
      grid-template-columns: repeat(6, minmax(0, 1fr));
      gap: 14px;
      margin: 18px 0 26px;
    }}
    .card {{
      background: linear-gradient(180deg, rgba(24, 35, 66, 0.95), rgba(15, 23, 48, 0.95));
      border: 1px solid rgba(103, 212, 255, 0.15);
      border-radius: 18px;
      padding: 16px;
    }}
    .label {{ color: var(--muted); font-size: 12px; margin-bottom: 6px; }}
    .value {{ font-size: 24px; font-weight: 700; }}
    .panel {{
      background: rgba(18, 26, 49, 0.96);
      border: 1px solid var(--line);
      border-radius: 24px;
      overflow: hidden;
    }}
    .panel-head {{
      display: flex;
      justify-content: space-between;
      gap: 20px;
      padding: 18px 22px;
      border-bottom: 1px solid var(--line);
      color: var(--muted);
    }}
    table {{ width: 100%; border-collapse: collapse; }}
    th, td {{
      padding: 14px 16px;
      border-bottom: 1px solid rgba(43, 59, 107, 0.55);
      vertical-align: top;
      text-align: left;
      font-size: 14px;
    }}
    th {{
      position: sticky;
      top: 0;
      background: rgba(24, 35, 66, 0.98);
      color: #d7e5ff;
      z-index: 1;
    }}
    tbody tr:nth-child(odd) {{ background: rgba(255, 255, 255, 0.01); }}
    tbody tr:hover {{ background: rgba(103, 212, 255, 0.06); }}
    .mono {{
      font-family: ui-monospace, SFMono-Regular, Menlo, monospace;
      font-size: 12px;
      word-break: break-all;
      color: #dce7ff;
    }}
    .foot {{
      margin-top: 16px;
      color: var(--muted);
      font-size: 13px;
      line-height: 1.6;
    }}
    @media (max-width: 1180px) {{ .grid {{ grid-template-columns: repeat(3, minmax(0, 1fr)); }} }}
    @media (max-width: 780px) {{
      body {{ padding: 18px; }}
      .grid {{ grid-template-columns: repeat(2, minmax(0, 1fr)); }}
      h1 {{ font-size: 28px; }}
      .panel {{ overflow: auto; }}
      table {{ min-width: 1100px; }}
    }}
  </style>
</head>
<body>
  <div class="wrap">
    <section class="hero">
      <div class="eyebrow">MarsChain Power Ranking</div>
      <h1>MarsChain 已发现地址算力榜</h1>
      <div class="sub">
        这份榜单基于 explorer 的公开接口生成，不是官方全网榜。
        它通过扫描近期区块、交易、POWER 合约日志、以及正算力地址的上级关系，拼出一份更接近全网的 best effort 排行。
      </div>
      <div class="grid">{summary_html}</div>
    </section>
    <section class="panel">
      <div class="panel-head">
        <div>Top {len(rows)} 地址</div>
        <div>扫描范围: {meta['tx_pages']} 页交易, {meta['block_pages']} 页区块, {meta.get('rpc_log_blocks_scanned', 0)} 个日志区块, upline 深度 {meta['upline_depth']}</div>
      </div>
      <table>
        <thead>
          <tr>
            <th>Rank</th>
            <th>Address</th>
            <th>Power</th>
            <th>Total Burned</th>
            <th>Tx Seen</th>
            <th>Miner Seen</th>
            <th>Log Seen</th>
            <th>Upline Seen</th>
            <th>Nodes</th>
            <th>Upline 1</th>
            <th>Upline 2</th>
          </tr>
        </thead>
        <tbody>
          {row_html}
        </tbody>
      </table>
    </section>
    <div class="foot">
      说明: 站点公开提供了单地址算力接口，但没有公开榜单接口。覆盖率表示当前已发现正算力地址的算力总和占 explorer 公布全网总算力的比例。
    </div>
  </div>
</body>
</html>
"""
    path.write_text(html)


def write_xlsx(path: Path, rows: list[RankedAddress], meta: dict[str, Any]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "排行榜"
    header = ["排名", "地址", "算力"]
    ws.append(header)
    for idx, row in enumerate(rows, start=1):
        ws.append([idx, row.address, row.power])
    fill = PatternFill("solid", fgColor="1F4B99")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {
        1: 8,
        2: 46,
        3: 22,
    }
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width
    wb.save(path)


def write_json(path: Path, rows: list[RankedAddress], meta: dict[str, Any]) -> None:
    path.write_text(
        json.dumps(
            {
                "meta": meta,
                "rows": [row.to_dict() for row in rows],
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n"
    )


def write_csv(path: Path, rows: list[RankedAddress]) -> None:
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(
            fh,
            fieldnames=["排名", "地址", "算力"],
        )
        writer.writeheader()
        for idx, row in enumerate(rows, start=1):
            writer.writerow({"排名": idx, "地址": row.address, "算力": row.power})


def lookup_power_rows(
    addresses: list[str],
    tx_counter: Counter[str],
    miner_counter: Counter[str],
    log_counter: Counter[str],
    upline_counter: Counter[str],
    args: argparse.Namespace,
    cache: dict[str, dict[str, Any]],
    progress_label: str,
) -> tuple[list[RankedAddress], dict[str, dict[str, Any]], Counter[str], dict[str, str]]:
    rows: list[RankedAddress] = []
    cache_updates: dict[str, dict[str, Any]] = {}
    cache_stats: Counter[str] = Counter()
    cache_status_by_address: dict[str, str] = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=args.workers) as pool:
        future_map = {
            pool.submit(fetch_power_payload, address, cache, args.cache_ttl_seconds): address
            for address in addresses
        }
        for idx, future in enumerate(concurrent.futures.as_completed(future_map), start=1):
            address = future_map[future]
            try:
                returned_address, payload, cache_status = future.result()
            except Exception as exc:
                print(f"[warn] power lookup failed for {address}: {exc}", file=sys.stderr)
                continue
            cache_stats[cache_status] += 1
            cache_status_by_address[returned_address] = cache_status
            cache_updates[returned_address] = payload
            row = build_row_from_payload(
                returned_address,
                payload,
                tx_seen=tx_counter[returned_address],
                miner_seen=miner_counter[returned_address],
                log_seen=log_counter[returned_address],
                upline_seen=upline_counter[returned_address],
            )
            if row:
                rows.append(row)
            if args.progress and idx % 100 == 0:
                print(f"[info] {progress_label}: checked {idx}/{len(addresses)} candidates", file=sys.stderr)
    return rows, cache_updates, cache_stats, cache_status_by_address


def build_rows_from_pool(
    pool: dict[str, dict[str, Any]],
    cache: dict[str, dict[str, Any]],
) -> list[RankedAddress]:
    rows: list[RankedAddress] = []
    for address, entry in pool.items():
        payload = cache.get(address)
        if not payload:
            continue
        row = build_row_from_payload(
            address,
            payload,
            tx_seen=int(entry.get("tx_seen_total", 0) or 0),
            miner_seen=int(entry.get("miner_seen_total", 0) or 0),
            log_seen=int(entry.get("log_seen_total", 0) or 0),
            upline_seen=int(entry.get("upline_seen_total", 0) or 0),
        )
        if row:
            rows.append(row)
    return rows


def select_refresh_candidates(
    pool: dict[str, dict[str, Any]],
    cache: dict[str, dict[str, Any]],
    current_scores: dict[str, int],
    max_candidates: int,
    cache_ttl_seconds: int | None,
    observed_at: int,
) -> list[str]:
    if max_candidates <= 0:
        budget = len(pool)
    else:
        budget = max_candidates
    if budget <= 0:
        return []

    ranked: list[tuple[tuple[int, int, int, int, str], str]] = []
    for address, entry in pool.items():
        cached = cache.get(address)
        if cached is not None and is_cache_entry_fresh(cached, cache_ttl_seconds):
            continue
        last_refresh_at = entry.get("last_network_refresh_at")
        if isinstance(last_refresh_at, int | float):
            refresh_age = max(0, observed_at - int(last_refresh_at))
        else:
            refresh_age = 10**12
        current_score = int(current_scores.get(address, 0) or 0)
        pool_score = int(entry.get("source_score_total", 0) or 0)
        last_seen_at = int(entry.get("last_seen_at", 0) or 0)
        ranked.append(((refresh_age, current_score, pool_score, last_seen_at, address), address))

    ranked.sort(key=lambda item: item[0], reverse=True)
    return [address for _, address in ranked[:budget]]


def build_ranking(args: argparse.Namespace) -> tuple[list[RankedAddress], dict[str, Any]]:
    tx_counter, miner_counter = collect_candidates(
        tx_pages=args.tx_pages,
        tx_limit=args.tx_limit,
        block_pages=args.block_pages,
        block_limit=args.block_limit,
        include_to=args.include_to,
        workers=args.workers,
        progress=args.progress,
    )
    log_counter: Counter[str] = Counter()
    rpc_meta: dict[str, Any] = {}
    rpc_log_meta: dict[str, Any] = {}
    rpc_blocks = getattr(args, "rpc_blocks", 0)
    if rpc_blocks:
        rpc_tx_counter, rpc_miner_counter, rpc_meta = collect_rpc_block_candidates(
            rpc_url=getattr(args, "rpc_url", DEFAULT_RPC_URL),
            rpc_blocks=rpc_blocks,
            rpc_start_block=getattr(args, "rpc_start_block", None),
            rpc_batch_size=getattr(args, "rpc_batch_size", 100),
            rpc_workers=getattr(args, "rpc_workers", min(args.workers, 8)),
            include_to=args.include_to,
            progress=args.progress,
        )
        tx_counter.update(rpc_tx_counter)
        miner_counter.update(rpc_miner_counter)

    rpc_log_blocks = getattr(args, "rpc_log_blocks", 0)
    if rpc_log_blocks:
        log_counter, rpc_log_meta = collect_rpc_log_candidates(
            rpc_url=getattr(args, "rpc_url", DEFAULT_RPC_URL),
            rpc_log_blocks=rpc_log_blocks,
            rpc_log_start_block=getattr(args, "rpc_log_start_block", None),
            rpc_log_chunk_size=getattr(args, "rpc_log_chunk_size", 50_000),
            rpc_log_workers=getattr(args, "rpc_log_workers", 3),
            progress=args.progress,
        )

    now = int(time.time())
    current_scores = {
        address: tx_counter[address] * 10 + miner_counter[address] * 20 + log_counter[address] * 15
        for address in set(tx_counter) | set(miner_counter) | set(log_counter)
    }
    current_candidates = sorted(
        current_scores,
        key=lambda addr: (
            current_scores[addr],
            log_counter[addr],
            tx_counter[addr],
            miner_counter[addr],
            addr,
        ),
        reverse=True,
    )

    cache = load_cache(Path(args.cache_file) if args.cache_file else None)
    pool_path = Path(args.address_pool_file) if getattr(args, "address_pool_file", None) else None
    address_pool, address_pool_meta = load_address_pool(pool_path)
    bootstrap_complete = bool(address_pool_meta.get("bootstrap_complete"))
    add_cache_addresses_to_pool(address_pool, cache, now)
    for address in current_candidates:
        ensure_address_pool_entry(address_pool, address, now, "current")

    if bootstrap_complete:
        ordered_candidates = select_refresh_candidates(
            address_pool,
            cache,
            current_scores,
            args.max_candidates,
            args.cache_ttl_seconds,
            now,
        )
        refresh_mode = "incremental"
    else:
        ordered_candidates = sorted(
            address_pool,
            key=lambda addr: (
                current_scores.get(addr, 0),
                int(address_pool.get(addr, {}).get("source_score_total", 0) or 0),
                int(address_pool.get(addr, {}).get("last_seen_at", 0) or 0),
                addr,
            ),
            reverse=True,
        )
        refresh_mode = "bootstrap"
    bootstrap_complete_after = bootstrap_complete or bool(ordered_candidates)

    cache_lookup_stats: Counter[str] = Counter()
    upline_counter: Counter[str] = Counter()
    rows_map: dict[str, RankedAddress] = {}

    initial_rows, cache_updates, cache_stats, cache_status_by_address = lookup_power_rows(
        ordered_candidates,
        tx_counter,
        miner_counter,
        log_counter,
        upline_counter,
        args,
        cache,
        "initial",
    )
    cache_lookup_stats.update(cache_stats)
    cache.update(cache_updates)
    for row in initial_rows:
        rows_map[row.address] = row

    seen_addresses = set(ordered_candidates)
    for depth in range(1, args.upline_depth + 1):
        next_counter: Counter[str] = Counter()
        for row in rows_map.values():
            for candidate in (row.upline1, row.upline2):
                if is_probably_user_address(candidate) and candidate not in seen_addresses:
                    next_counter[candidate] += 1
        if not next_counter:
            break
        next_addresses = [address for address, _ in next_counter.most_common(args.upline_limit)]
        if not next_addresses:
            break
        seen_addresses.update(next_addresses)
        upline_counter.update(next_counter)
        depth_rows, cache_updates, cache_stats, depth_cache_status = lookup_power_rows(
            next_addresses,
            tx_counter,
            miner_counter,
            log_counter,
            upline_counter,
            args,
            cache,
            f"upline-depth-{depth}",
        )
        cache_lookup_stats.update(cache_stats)
        cache.update(cache_updates)
        cache_status_by_address.update(depth_cache_status)
        for row in depth_rows:
            rows_map[row.address] = row

    expanded_history_addresses: set[str] = set()
    for depth in range(1, args.history_depth + 1):
        seed_rows = [
            row
            for row in sorted(rows_map.values(), key=lambda row: (row.power, row.total_burned_amount), reverse=True)
            if row.address not in expanded_history_addresses
        ][: args.history_seed_limit]
        if not seed_rows:
            break
        seed_addresses = [row.address for row in seed_rows]
        expanded_history_addresses.update(seed_addresses)

        history_counter = collect_history_counter(
            seed_addresses=seed_addresses,
            pages=args.history_pages,
            limit=args.history_tx_limit,
            workers=args.workers,
            progress=args.progress,
        )
        if not history_counter:
            break

        next_addresses = [
            address
            for address, _ in history_counter.most_common(args.history_candidate_limit)
            if address not in seen_addresses
        ]
        if not next_addresses:
            continue

        seen_addresses.update(next_addresses)
        tx_counter.update(history_counter)
        depth_rows, cache_updates, cache_stats, depth_cache_status = lookup_power_rows(
            next_addresses,
            tx_counter,
            miner_counter,
            log_counter,
            upline_counter,
            args,
            cache,
            f"history-depth-{depth}",
        )
        cache_lookup_stats.update(cache_stats)
        cache.update(cache_updates)
        cache_status_by_address.update(depth_cache_status)
        for row in depth_rows:
            rows_map[row.address] = row

    for address, status in cache_status_by_address.items():
        if status == "refreshed":
            mark_network_refresh(address_pool, address, now, "power")

    merge_address_pool_counters(
        address_pool,
        tx_counter,
        miner_counter,
        log_counter,
        upline_counter,
        now,
    )
    save_cache(Path(args.cache_file) if args.cache_file else None, cache)
    pool_meta = {
        "schema_version": 1,
        "generated_at": now,
        "generated_at_local": format_beijing_datetime(now),
        "address_count": len(address_pool),
        "cached_address_count": len(cache),
        "selected_refresh_count": len(ordered_candidates),
        "current_candidate_count": len(current_candidates),
        "source_score_total": sum(int(entry.get("source_score_total", 0) or 0) for entry in address_pool.values()),
        "bootstrap_complete": bootstrap_complete_after,
        "refresh_mode": refresh_mode,
    }
    address_pool_meta.update(pool_meta)
    save_address_pool(pool_path, address_pool, address_pool_meta)

    all_rows = build_rows_from_pool(address_pool, cache)
    all_rows.sort(
        key=lambda row: (
            row.power,
            row.total_burned_amount,
            row.source_score,
            row.tx_seen,
            row.miner_seen,
            row.upline_seen,
            row.address,
        ),
        reverse=True,
    )
    rows = all_rows if args.top <= 0 else all_rows[: args.top]

    if args.include_nodes and rows:
        with concurrent.futures.ThreadPoolExecutor(max_workers=min(args.workers, 8)) as pool:
            future_map = {pool.submit(enrich_nodes, row.address): row for row in rows}
            for future in concurrent.futures.as_completed(future_map):
                row = future_map[future]
                try:
                    row.nodes_count = future.result()
                except Exception as exc:
                    print(f"[warn] nodes lookup failed for {row.address}: {exc}", file=sys.stderr)

    try:
        network_stats = request_json("/stats")
    except Exception as exc:
        network_stats = {}
        print(f"[warn] network stats lookup failed: {exc}", file=sys.stderr)

    power_stats = request_json("/power/stats")
    network_total_power = int(power_stats.get("totalPower", "0") or 0)
    network_total_burned_tokens = int(power_stats.get("totalBurnedTokens", "0") or 0)
    explorer_total_addresses = int(network_stats.get("totalAddresses", 0) or 0) if isinstance(network_stats, dict) else 0
    network_total_circulation_tokens = int(network_stats.get("totalCirculation", "0") or 0) if isinstance(network_stats, dict) else 0
    network_current_price = network_stats.get("currentPrice") if isinstance(network_stats, dict) else None
    network_highest_price = network_stats.get("highestPrice") if isinstance(network_stats, dict) else None
    network_lowest_price = network_stats.get("lowestPrice") if isinstance(network_stats, dict) else None
    latest_block = rpc_log_meta.get("rpc_log_latest_block") if rpc_log_meta else None
    if latest_block is None:
        latest_block = rpc_meta.get("rpc_latest_block") if rpc_meta else None
    statistics_reference_timestamp = rpc_log_meta.get("rpc_log_latest_timestamp") if rpc_log_meta else None
    if not isinstance(statistics_reference_timestamp, int):
        statistics_reference_timestamp = int(time.time())
    statistics_window_meta = build_statistics_window_meta(statistics_reference_timestamp)
    emission_meta = build_mars_emission_meta(args.rpc_url, statistics_reference_timestamp, network_total_power)
    statistics_window_active_meta = collect_statistics_window_active_addresses(
        args.rpc_url,
        rpc_log_meta.get("rpc_log_latest_block") if rpc_log_meta else None,
        statistics_window_meta["statistics_window_start_timestamp"],
        statistics_window_meta["statistics_window_end_timestamp"],
        args.rpc_batch_size,
        args.rpc_workers,
        args.progress,
    )

    daily_power_history_days = 0
    statistics_window_start_day = None
    statistics_window_end_day = None
    today_chain_day = None
    today_utc_date = None
    today_local_date = statistics_window_meta["statistics_day_label"]
    today_new_power = None
    statistics_window_start_total_power = None
    statistics_window_end_total_power = None
    period_power_meta: dict[str, Any] = {}
    daily_total_power_history: list[dict[str, Any]] = []
    try:
        history_days, history_powers = fetch_daily_total_power_history(args.rpc_url)
        daily_power_history_days = len(history_days)
        history_map = dict(zip(history_days, history_powers))
        daily_total_power_history = [
            {
                "day": int(day),
                "date": time.strftime("%Y-%m-%d", time.gmtime(int(day) * 86_400)),
                "value": int(power),
            }
            for day, power in sorted(history_map.items())[-90:]
        ]

        def lookup_total_power_for_day(day: int) -> int | None:
            if day in history_map:
                return history_map[day]
            previous_days = [history_day for history_day in history_days if history_day < day]
            if previous_days:
                return history_map[max(previous_days)]
            return None

        statistics_window_start_day = int(statistics_window_meta["statistics_window_start_timestamp"] // 86_400)
        statistics_window_end_day = int(statistics_window_meta["statistics_window_end_timestamp"] // 86_400)
        today_chain_day = statistics_window_start_day
        today_utc_date = time.strftime("%Y-%m-%d", time.gmtime(statistics_window_start_day * 86_400))
        statistics_window_start_total_power = lookup_total_power_for_day(statistics_window_start_day)
        statistics_window_end_total_power = history_map.get(statistics_window_end_day)
        if statistics_window_end_total_power is None:
            statistics_window_end_total_power = network_total_power
        if statistics_window_end_day is not None and statistics_window_end_total_power is not None:
            if not daily_total_power_history or daily_total_power_history[-1]["day"] != statistics_window_end_day:
                daily_total_power_history.append(
                    {
                        "day": int(statistics_window_end_day),
                        "date": time.strftime("%Y-%m-%d", time.gmtime(int(statistics_window_end_day) * 86_400)),
                        "value": int(statistics_window_end_total_power),
                    }
                )
                daily_total_power_history = daily_total_power_history[-90:]
        if statistics_window_start_total_power is not None and statistics_window_end_total_power is not None:
            today_new_power = max(0, statistics_window_end_total_power - statistics_window_start_total_power)
        for days in (7, 30):
            prefix = f"period_{days}d"
            period_end_day = statistics_window_end_day
            period_start_day = max(0, period_end_day - days)
            period_start_total_power = lookup_total_power_for_day(period_start_day)
            period_end_total_power = history_map.get(period_end_day)
            if period_end_total_power is None:
                period_end_total_power = network_total_power
            period_new_power = None
            if period_start_total_power is not None and period_end_total_power is not None:
                period_new_power = max(0, period_end_total_power - period_start_total_power)
            period_power_meta.update(
                {
                    f"{prefix}_start_day": period_start_day,
                    f"{prefix}_end_day": period_end_day,
                    f"{prefix}_start_total_power": period_start_total_power,
                    f"{prefix}_end_total_power": period_end_total_power,
                    f"{prefix}_new_power": period_new_power,
                    f"{prefix}_new_power_basis": f"latest completed {days} Beijing statistics days end totalPower minus start totalPower",
                }
            )
    except Exception as exc:
        print(f"[warn] daily total power history lookup failed: {exc}", file=sys.stderr)

    discovered_total_power = sum(row.power for row in all_rows)
    generated_at = int(time.time())
    meta = {
        "generated_at": generated_at,
        "generated_at_local": format_beijing_datetime(generated_at),
        "base_url": BASE_URL,
        "ranking_type": "best_effort_discovered_addresses",
        "tx_pages": args.tx_pages,
        "tx_limit": args.tx_limit,
        "block_pages": args.block_pages,
        "block_limit": args.block_limit,
        "rpc_enabled": bool(rpc_blocks),
        "rpc_url": rpc_meta.get("rpc_url") if rpc_meta else None,
        "rpc_blocks_requested": rpc_meta.get("rpc_blocks_requested", 0),
        "rpc_blocks_scanned": rpc_meta.get("rpc_blocks_scanned", 0),
        "rpc_transactions_seen": rpc_meta.get("rpc_transactions_seen", 0),
        "rpc_start_block": rpc_meta.get("rpc_start_block"),
        "rpc_end_block": rpc_meta.get("rpc_end_block"),
        "rpc_latest_block": rpc_meta.get("rpc_latest_block"),
        "rpc_error_count": rpc_meta.get("rpc_error_count", 0),
        "rpc_log_enabled": bool(rpc_log_blocks),
        "rpc_log_contract": rpc_log_meta.get("rpc_log_contract") if rpc_log_meta else None,
        "rpc_log_blocks_requested": rpc_log_meta.get("rpc_log_blocks_requested", 0),
        "rpc_log_blocks_effective": rpc_log_meta.get("rpc_log_blocks_effective", 0),
        "rpc_log_blocks_scanned": rpc_log_meta.get("rpc_log_blocks_scanned", 0),
        "rpc_logs_seen": rpc_log_meta.get("rpc_logs_seen", 0),
        "rpc_log_addresses_seen": rpc_log_meta.get("rpc_log_addresses_seen", 0),
        "rpc_log_first_seen_tracked": rpc_log_meta.get("rpc_log_first_seen_tracked", 0),
        "rpc_log_start_block": rpc_log_meta.get("rpc_log_start_block"),
        "rpc_log_end_block": rpc_log_meta.get("rpc_log_end_block"),
        "rpc_log_latest_block": rpc_log_meta.get("rpc_log_latest_block"),
        "rpc_log_latest_timestamp": rpc_log_meta.get("rpc_log_latest_timestamp"),
        "rpc_log_today_start_timestamp": rpc_log_meta.get("rpc_log_today_start_timestamp"),
        "rpc_log_today_start_block": rpc_log_meta.get("rpc_log_today_start_block"),
        "rpc_log_statistics_window_start_timestamp": rpc_log_meta.get("rpc_log_statistics_window_start_timestamp"),
        "rpc_log_statistics_window_end_timestamp": rpc_log_meta.get("rpc_log_statistics_window_end_timestamp"),
        "rpc_log_statistics_window_start_block": rpc_log_meta.get("rpc_log_statistics_window_start_block"),
        "rpc_log_statistics_window_end_block": rpc_log_meta.get("rpc_log_statistics_window_end_block"),
        "rpc_log_error_count": rpc_log_meta.get("rpc_log_error_count", 0),
        **{
            key: rpc_log_meta.get(key)
            for days in (7, 30)
            for key in (
                f"period_{days}d_label",
                f"period_{days}d_start_timestamp",
                f"period_{days}d_end_timestamp",
                f"period_{days}d_start_local",
                f"period_{days}d_end_local",
                f"period_{days}d_start_block",
                f"period_{days}d_end_block",
                f"period_{days}d_complete",
                f"period_{days}d_new_candidate_address_count",
                f"period_{days}d_new_candidate_address_basis",
                f"period_{days}d_burned_tokens",
                f"period_{days}d_burned_display",
                f"period_{days}d_burned_basis",
            )
        },
        **statistics_window_meta,
        **emission_meta,
        **statistics_window_active_meta,
        **period_power_meta,
        "statistics_window_start_day": statistics_window_start_day,
        "statistics_window_end_day": statistics_window_end_day,
        "today_chain_day": today_chain_day,
        "today_utc_date": today_utc_date,
        "today_local_date": today_local_date,
        "today_new_wallet_count": rpc_log_meta.get("today_new_wallet_count"),
        "today_new_wallet_basis": rpc_log_meta.get("today_new_wallet_basis"),
        "statistics_window_new_candidate_address_count": rpc_log_meta.get("statistics_window_new_candidate_address_count"),
        "statistics_window_new_candidate_address_basis": rpc_log_meta.get("statistics_window_new_candidate_address_basis"),
        "statistics_window_burned_tokens": rpc_log_meta.get("statistics_window_burned_tokens"),
        "statistics_window_burned_display": rpc_log_meta.get("statistics_window_burned_display"),
        "statistics_window_burned_basis": rpc_log_meta.get("statistics_window_burned_basis"),
        "today_burned_tokens": rpc_log_meta.get("today_burned_tokens"),
        "today_burned_display": rpc_log_meta.get("today_burned_display"),
        "today_burned_basis": rpc_log_meta.get("today_burned_basis"),
        "today_new_power": today_new_power,
        "statistics_window_new_power": today_new_power,
        "today_new_power_basis": f"completed Beijing {STATISTICS_DAY_START_LABEL} statistics window end totalPower minus start totalPower",
        "statistics_window_new_power_basis": f"completed Beijing {STATISTICS_DAY_START_LABEL} statistics window end totalPower minus start totalPower",
        "previous_day_total_power": statistics_window_start_total_power,
        "statistics_window_start_total_power": statistics_window_start_total_power,
        "statistics_window_end_total_power": statistics_window_end_total_power,
        "daily_power_history_days": daily_power_history_days,
        "daily_total_power_history": daily_total_power_history,
        "include_to": args.include_to,
        "upline_depth": args.upline_depth,
        "upline_limit": args.upline_limit,
        "history_depth": args.history_depth,
        "history_pages": args.history_pages,
        "history_tx_limit": args.history_tx_limit,
        "history_seed_limit": args.history_seed_limit,
        "history_candidate_limit": args.history_candidate_limit,
        "power_cache_ttl_seconds": args.cache_ttl_seconds,
        "power_cache_hits": cache_lookup_stats.get("hit", 0),
        "power_cache_refreshed": cache_lookup_stats.get("refreshed", 0),
        "power_cache_stale_fallbacks": cache_lookup_stats.get("stale_fallback", 0),
        "seed_candidate_count": len(ordered_candidates),
        "candidate_count": len(address_pool),
        "address_pool_count": len(address_pool),
        "bootstrap_complete": bootstrap_complete_after,
        "refresh_mode": refresh_mode,
        "address_pool_bootstrap_complete": bootstrap_complete_after,
        "address_pool_refresh_mode": refresh_mode,
        "positive_power_count": len(all_rows),
        "explorer_total_addresses": explorer_total_addresses,
        "network_total_circulation_tokens": network_total_circulation_tokens,
        "network_total_circulation_display": format_token_chinese(network_total_circulation_tokens),
        "network_current_price": network_current_price,
        "network_current_price_display": format_price(network_current_price),
        "network_highest_price": network_highest_price,
        "network_highest_price_display": format_price(network_highest_price),
        "network_lowest_price": network_lowest_price,
        "network_lowest_price_display": format_price(network_lowest_price),
        "discovered_total_power": discovered_total_power,
        "network_total_power": network_total_power,
        "network_total_burned_tokens": network_total_burned_tokens,
        "network_total_burned_display": format_token_chinese(network_total_burned_tokens),
        "latest_block": latest_block,
        "discovered_power_coverage": (discovered_total_power / network_total_power) if network_total_power else 0.0,
        "ranked_count": len(rows),
    }
    return rows, meta


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build a best-effort MarsChain power ranking.")
    parser.add_argument("--tx-pages", type=int, default=60, help="Recent transaction pages to scan.")
    parser.add_argument("--tx-limit", type=int, default=100, help="Transactions per page.")
    parser.add_argument("--block-pages", type=int, default=20, help="Recent block pages to scan.")
    parser.add_argument("--block-limit", type=int, default=100, help="Blocks per page.")
    parser.add_argument("--rpc-url", default=DEFAULT_RPC_URL, help="Optional JSON-RPC endpoint for extra block/tx candidate scanning.")
    parser.add_argument("--rpc-blocks", type=int, default=0, help="Extra recent blocks to scan through JSON-RPC.")
    parser.add_argument("--rpc-start-block", type=int, default=None, help="Highest block number for JSON-RPC scanning; defaults to latest.")
    parser.add_argument("--rpc-batch-size", type=int, default=100, help="Blocks per JSON-RPC batch request.")
    parser.add_argument("--rpc-workers", type=int, default=6, help="Concurrent JSON-RPC batch workers.")
    parser.add_argument("--rpc-log-blocks", type=int, default=0, help="Recent blocks to scan for POWER contract logs through JSON-RPC.")
    parser.add_argument("--rpc-log-start-block", type=int, default=None, help="Highest block number for JSON-RPC log scanning; defaults to latest.")
    parser.add_argument("--rpc-log-chunk-size", type=int, default=50_000, help="Blocks per eth_getLogs range.")
    parser.add_argument("--rpc-log-workers", type=int, default=3, help="Concurrent eth_getLogs workers.")
    parser.add_argument("--max-candidates", type=int, default=3000, help="Maximum candidate addresses to power-check.")
    parser.add_argument("--top", type=int, default=100, help="Number of ranked rows to output. Use 0 to output every discovered positive-power row.")
    parser.add_argument("--workers", type=int, default=16, help="Concurrent power lookups.")
    parser.add_argument("--include-to", action="store_true", help="Also include transaction recipient addresses as candidates.")
    parser.add_argument("--include-nodes", action="store_true", help="Fetch /nodes/{address} for final ranked rows.")
    parser.add_argument("--upline-depth", type=int, default=2, help="How many rounds of discovered uplines to follow.")
    parser.add_argument("--upline-limit", type=int, default=2500, help="Maximum upline candidates per depth.")
    parser.add_argument("--history-depth", type=int, default=0, help="How many rounds of address-history expansion to run.")
    parser.add_argument("--history-pages", type=int, default=2, help="Transaction pages to scan per history seed address.")
    parser.add_argument("--history-tx-limit", type=int, default=100, help="Transactions per page for address-history expansion.")
    parser.add_argument("--history-seed-limit", type=int, default=200, help="How many top power addresses to expand per history round.")
    parser.add_argument("--history-candidate-limit", type=int, default=50000, help="Maximum counterparties to power-check per history round.")
    parser.add_argument("--output-dir", default="output", help="Directory for JSON and CSV outputs.")
    parser.add_argument("--prefix", default="marschain_power_rank", help="Output filename prefix.")
    parser.add_argument("--cache-file", default="output/marschain_power_cache.json", help="Power lookup cache JSON path.")
    parser.add_argument("--address-pool-file", default=DEFAULT_ADDRESS_POOL_FILE, help="Persistent discovered address pool JSON path.")
    parser.add_argument(
        "--cache-ttl-seconds",
        type=int,
        default=DEFAULT_CACHE_TTL_SECONDS,
        help="Power lookup cache TTL. Use 0 to force refresh; use -1 to trust cache forever.",
    )
    parser.add_argument("--progress", action="store_true", help="Print progress to stderr.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    rows, meta = build_ranking(args)

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    stamp = time.strftime("%Y%m%d_%H%M%S")
    json_path = output_dir / f"{args.prefix}_{stamp}.json"
    csv_path = output_dir / f"{args.prefix}_{stamp}.csv"
    html_path = output_dir / f"{args.prefix}_{stamp}.html"
    xlsx_path = output_dir / f"{args.prefix}_{stamp}.xlsx"
    write_json(json_path, rows, meta)
    write_csv(csv_path, rows)
    write_html(html_path, rows, meta)
    write_xlsx(xlsx_path, rows, meta)

    print(f"Generated {len(rows)} rows from {meta['candidate_count']} candidates.")
    print(f"JSON: {json_path}")
    print(f"CSV:  {csv_path}")
    print(f"HTML: {html_path}")
    print(f"XLSX: {xlsx_path}")
    if rows:
        print("\nTop 10")
        for idx, row in enumerate(rows[:10], start=1):
            print(
                f"{idx:>2}. {row.address}  power={row.power} ({format_units(row.power)})  "
                f"burned={format_token(row.total_burned_amount)}  tx_seen={row.tx_seen} "
                f"miner_seen={row.miner_seen} upline_seen={row.upline_seen}"
            )
    else:
        print("No ranked rows found. Try scanning more tx pages or using --include-to.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
